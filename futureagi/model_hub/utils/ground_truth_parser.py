"""
Ground truth file parser — handles CSV, Excel (XLS/XLSX), and JSON uploads.
"""

import csv
import io
import json

import structlog

logger = structlog.get_logger(__name__)

MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024  # 50 MB


def parse_ground_truth_file(file_obj, file_name: str) -> tuple[list[str], list[dict]]:
    """
    Parse an uploaded file into columns and row data.

    Args:
        file_obj: Django UploadedFile or file-like object
        file_name: Original file name (used to determine format)

    Returns:
        (columns, data) — list of column names and list of row dicts

    Raises:
        ValueError: If file format is unsupported, file is empty, or parsing fails
    """
    ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""

    if ext == "json":
        return _parse_json(file_obj)
    elif ext == "csv":
        return _parse_csv(file_obj)
    elif ext in ("xls", "xlsx"):
        return _parse_excel(file_obj)
    else:
        raise ValueError(
            f"Unsupported file type: '.{ext}'. Supported: CSV, XLS, XLSX, JSON."
        )


def _parse_csv(file_obj) -> tuple[list[str], list[dict]]:
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")  # Handle BOM

    reader = csv.DictReader(io.StringIO(content))
    columns = reader.fieldnames
    if not columns:
        raise ValueError("CSV file has no headers.")

    columns = [c.lstrip("\ufeff").strip() for c in columns]
    data = []
    for row in reader:
        cleaned = {k.lstrip("\ufeff").strip(): v for k, v in row.items() if k}
        data.append(cleaned)

    if not data:
        raise ValueError("CSV file has headers but no data rows.")

    return columns, data


def _parse_json(file_obj) -> tuple[list[str], list[dict]]:
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8")

    try:
        parsed = json.loads(content)
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON: {e}")

    if isinstance(parsed, list):
        if not parsed:
            raise ValueError("JSON array is empty.")
        if not isinstance(parsed[0], dict):
            raise ValueError("JSON array must contain objects (dicts), not primitives.")
        columns = list(parsed[0].keys())
        return columns, parsed

    if isinstance(parsed, dict):
        # Support {"columns": [...], "data": [...]} format
        if "columns" in parsed and "data" in parsed:
            return parsed["columns"], parsed["data"]
        # Single object — wrap in list
        columns = list(parsed.keys())
        return columns, [parsed]

    raise ValueError("JSON must be an array of objects or {columns, data} dict.")


def _parse_excel(file_obj) -> tuple[list[str], list[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError(
            "Excel parsing requires openpyxl. Install it with: pip install openpyxl"
        )

    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to read Excel file: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Excel file is empty.")

    if len(rows) < 2:
        raise ValueError("Excel file has headers but no data rows.")

    columns = [
        str(c).strip() if c is not None else f"column_{i}"
        for i, c in enumerate(rows[0])
    ]

    data = []
    for row in rows[1:]:
        row_dict = {}
        for col_name, value in zip(columns, row):
            # Convert Excel types to JSON-safe values
            if value is None:
                row_dict[col_name] = ""
            else:
                row_dict[col_name] = (
                    str(value) if not isinstance(value, (int, float, bool)) else value
                )
        data.append(row_dict)

    return columns, data
